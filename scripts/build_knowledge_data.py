#!/usr/bin/env python3
"""Build data.js from V2 knowledge-base CSV files or workbook sheets."""

from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Iterable


SKIP_SHEETS = {"📋 总览", "总览", "字段说明", "分类统计", "知识库_All", "App导入_精简版"}

FIELD_ALIASES = {
    "term": ("Term", "名词/术语", "名词 / 术语", "名词/缩写", "名词 / 缩写", "术语", "名词"),
    "short": ("ShortDesc", "一句话解释", "一句话说明", "解释"),
    "detail": ("DetailedDesc", "详细大白话解释", "详细解释", "详细解释（指标类含公式）"),
    "formula": ("Formula", "公式", "公式/计算方式", "计算方式"),
    "example": ("Example", "场景", "例子/使用场景", "使用场景"),
    "confusing": ("Confusing", "易混淆", "易混淆概念"),
    "english": ("EnglishFull", "英文全称"),
    "related": ("RelatedTerms", "相关术语"),
    "mastery": ("Mastery", "默认掌握度"),
}


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def slugify(value: str) -> str:
    slug = re.sub(r"[\s/\\]+", "_", value.strip().lower())
    slug = re.sub(r"[^\w\u4e00-\u9fff.-]+", "", slug)
    return slug.strip("._-") or "term"


def split_terms(value: str) -> list[str]:
    return [part.strip() for part in re.split(r"[,，;；、]\s*", value or "") if part.strip()]


def pick(row: dict[str, object], aliases: Iterable[str]) -> str:
    normalized = {clean(key).lower().replace(" ", ""): value for key, value in row.items()}
    for alias in aliases:
        key = clean(alias).lower().replace(" ", "")
        if key in normalized and clean(normalized[key]):
            return clean(normalized[key])
    return ""


def find_header_index(rows: list[list[object]]) -> int:
    for index, row in enumerate(rows[:8]):
        labels = {clean(cell).lower().replace(" ", "") for cell in row}
        has_term = any(clean(alias).lower().replace(" ", "") in labels for alias in FIELD_ALIASES["term"])
        has_desc = any(clean(alias).lower().replace(" ", "") in labels for alias in FIELD_ALIASES["short"])
        if has_term and has_desc:
            return index
    raise ValueError("Could not find a V2 header row with term and short description columns.")


def rows_to_items(rows: list[list[object]], category: str, source_name: str) -> list[dict[str, object]]:
    header_index = find_header_index(rows)
    headers = [clean(cell) for cell in rows[header_index]]
    items: list[dict[str, object]] = []
    for row_index, raw_row in enumerate(rows[header_index + 1 :], start=1):
        row = {headers[index]: raw_row[index] if index < len(raw_row) else "" for index in range(len(headers))}
        term = pick(row, FIELD_ALIASES["term"])
        short_desc = pick(row, FIELD_ALIASES["short"])
        if not term or not short_desc:
            continue
        detail = pick(row, FIELD_ALIASES["detail"])
        mastery_raw = pick(row, FIELD_ALIASES["mastery"])
        try:
            mastery = max(1, min(5, int(float(mastery_raw)))) if mastery_raw else 3
        except ValueError:
            mastery = 3
        item_id = f"{slugify(category)}-{slugify(term)}-{row_index}"
        example = pick(row, FIELD_ALIASES["example"])
        items.append(
            {
                "id": item_id,
                "Term": term,
                "ShortDesc": short_desc,
                "DetailedDesc": detail,
                "Formula": pick(row, FIELD_ALIASES["formula"]),
                "Example": example,
                "Confusing": pick(row, FIELD_ALIASES["confusing"]),
                "EnglishFull": pick(row, FIELD_ALIASES["english"]),
                "Category": category,
                "Source": source_name,
                "term": term,
                "category": category,
                "definition": short_desc,
                "summary": short_desc,
                "detail": detail,
                "formula": pick(row, FIELD_ALIASES["formula"]),
                "examples": [example] if example else [],
                "confusion": pick(row, FIELD_ALIASES["confusing"]),
                "english": pick(row, FIELD_ALIASES["english"]),
                "relatedTerms": split_terms(pick(row, FIELD_ALIASES["related"])),
                "mastery": mastery,
                "questions": [],
            }
        )
    return items


def read_csv(path: Path) -> list[list[object]]:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            with path.open(newline="", encoding=encoding) as handle:
                return list(csv.reader(handle))
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("csv", b"", 0, 1, f"Unable to decode {path}")


def iter_sources(inputs: list[Path]) -> Iterable[tuple[str, str, list[list[object]]]]:
    for input_path in inputs:
        if input_path.is_dir():
            for csv_path in sorted(input_path.glob("*.csv")):
                yield csv_path.stem, csv_path.name, read_csv(csv_path)
            continue
        if input_path.suffix.lower() == ".csv":
            yield input_path.stem, input_path.name, read_csv(input_path)
            continue
        if input_path.suffix.lower() in {".xlsx", ".xlsm"}:
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise RuntimeError("openpyxl is required for workbook imports; use CSV inputs or install openpyxl.") from exc
            workbook = load_workbook(input_path, read_only=True, data_only=True)
            for sheet in workbook.worksheets:
                category = clean(sheet.title)
                if category in SKIP_SHEETS:
                    continue
                rows = [list(row) for row in sheet.iter_rows(values_only=True) if any(cell is not None for cell in row)]
                if rows:
                    yield category, f"{input_path.name}:{category}", rows
            continue
        raise ValueError(f"Unsupported input: {input_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate app data.js from V2 knowledge-base sources.")
    parser.add_argument("inputs", nargs="+", type=Path, help="CSV files, folders of CSVs, or .xlsx workbooks.")
    parser.add_argument("--out", type=Path, default=Path("data.js"), help="Output JavaScript file.")
    args = parser.parse_args()

    items: list[dict[str, object]] = []
    for category, source_name, rows in iter_sources(args.inputs):
        items.extend(rows_to_items(rows, category, source_name))

    seen: set[str] = set()
    for index, item in enumerate(items, start=1):
        base_id = clean(item["id"])
        item_id = base_id
        suffix = 2
        while item_id in seen:
            item_id = f"{base_id}-{suffix}"
            suffix += 1
        seen.add(item_id)
        item["id"] = item_id

    content = "// Generated by scripts/build_knowledge_data.py from V2 knowledge-base sources.\n"
    content += f"// Total items: {len(items)}.\n"
    content += "window.KNOWLEDGE_ITEMS = "
    content += json.dumps(items, ensure_ascii=False, indent=2)
    content += ";\n"
    args.out.write_text(content, encoding="utf-8")
    print(f"Wrote {len(items)} items to {args.out}")


if __name__ == "__main__":
    main()
